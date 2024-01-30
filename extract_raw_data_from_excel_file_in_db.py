"""
    Скрипт имитирует работу коннектора программы 1С для экспорта
    данных в БД.
    На данном этапе данные из программы 1С, вручную
    копируются в файл excel на отдельные листы, предназначенные
    для каждого вида отчета.
    Скрипт последовательно проходит по листам excel и собирает
    данные по заказу на производство в словарь.
    После сбора всей информации с листов, скрипт импортирует
    данные из словаря в таблицы БД.
"""

from pprint import pprint

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from app import (ArticleExtractor, Base, ReportSettingsOrders,
                 extract_data_job_monitor_for_work_centers,
                 extract_data_moving_sets_of_furniture,
                 extract_data_release_of_assembly_kits,
                 import_data_to_db_production_orders)

config = ReportSettingsOrders()
extractor = ArticleExtractor()
orders_data = dict()
error_log = dict()

if __name__ == '__main__':
    # Открыть файл с исходными данными
    workbook = load_workbook(
        filename=config.source_file,
        read_only=True,
        data_only=True
    )

    # Собираем данные по перемещению комплектов мебели
    orders_data = extract_data_moving_sets_of_furniture(
        orders_data=orders_data,
        error_log=error_log,
        workbook=workbook,
        sheet=config.sheet_moving_1C,
        expression=config.expression,
        extractor=extractor,
        config=config
    )

    # Собираем данные по выпуску комплектов мебели
    orders_data = extract_data_release_of_assembly_kits(
        orders_data=orders_data,
        workbook=workbook,
        sheet=config.sheet_kits_1C,
        expression=config.expression
    )

    # Собираем данные по проценту готовности заказов
    orders_data = extract_data_job_monitor_for_work_centers(
        orders_data=orders_data,
        workbook=workbook,
        sheet=config.sheet_percentage_1C
    )

    workbook_sheet = workbook[config.sheet_division_1C]
    child_orders_temp = dict()

    for value in workbook_sheet.iter_rows(min_row=2, values_only=True):
        full_order_date = value[0]
        order_date = full_order_date[6:10]
        order_number = "{:05d}".format(value[1])
        composite_key = order_date + order_number
        order_company = value[2]
        order_division = value[3]
        order_launch_date = value[4]
        order_execution_date = value[5]
        responsible = value[6]
        comment = value[7]
        order_parent = value[8]
        if order_parent:
            composite_key_parent = order_parent[43:47] + order_parent[28:33]
            if orders_data.get(composite_key_parent) and not orders_data[composite_key_parent].get('child_orders'):
                child_orders_temp[composite_key] = composite_key_parent
                orders_data[composite_key_parent]['child_orders'] = {
                    'report description of production orders': {
                        composite_key: {
                            'order_date': order_date,
                            'order_number': order_number,
                            'order_company': order_company,
                            'order_division': order_division,
                            'order_launch_date': order_launch_date,
                            'order_execution_date': order_execution_date,
                            'responsible': responsible,
                            'comment': comment,
                            }
                            }
                            }
            if orders_data.get(composite_key_parent) and orders_data[composite_key_parent]['child_orders'].get('report description of production orders'):
                child_orders_temp[composite_key] = composite_key_parent
                orders_data[composite_key_parent]['child_orders']['report description of production orders'][composite_key] = {
                    'order_date': order_date,
                    'order_number': order_number,
                    'order_company': order_company,
                    'order_division': order_division,
                    'order_launch_date': order_launch_date,
                    'order_execution_date': order_execution_date,
                    'responsible': responsible,
                    'comment': comment,
                    }
            if orders_data.get(composite_key_parent) and not orders_data[composite_key_parent]['child_orders'].get('report description of production orders'):
                child_orders_temp[composite_key] = composite_key_parent
                orders_data[composite_key_parent]['child_orders']['report description of production orders'] = {
                    composite_key: {
                        'order_date': order_date,
                        'order_number': order_number,
                        'order_company': order_company,
                        'order_division': order_division,
                        'order_launch_date': order_launch_date,
                        'order_execution_date': order_execution_date,
                        'responsible': responsible,
                        'comment': comment,
                        }
                }
            if child_orders_temp.get(composite_key_parent) and orders_data.get(child_orders_temp[composite_key_parent]) and orders_data[child_orders_temp[composite_key_parent]]['child_orders'].get('report description of production orders'):
                orders_data[child_orders_temp[composite_key_parent]]['child_orders']['report description of production orders'][composite_key] = {
                    'order_date': order_date,
                    'order_number': order_number,
                    'order_company': order_company,
                    'order_division': order_division,
                    'order_launch_date': order_launch_date,
                    'order_execution_date': order_execution_date,
                    'responsible': responsible,
                    'comment': comment,
                    }
            if child_orders_temp.get(composite_key_parent) and orders_data.get(child_orders_temp[composite_key_parent]) and not orders_data[child_orders_temp[composite_key_parent]]['child_orders'].get('report description of production orders'):
                child_orders_temp[composite_key] = composite_key_parent
                orders_data[child_orders_temp[composite_key_parent]]['child_orders']['report description of production orders'] = {
                    composite_key: {
                        'order_date': order_date,
                        'order_number': order_number,
                        'order_company': order_company,
                        'order_division': order_division,
                        'order_launch_date': order_launch_date,
                        'order_execution_date': order_execution_date,
                        'responsible': responsible,
                        'comment': comment,
                        }
                }
        else:
            if orders_data.get(composite_key):
                orders_data[composite_key]['report description of production orders'] = {
                        'order_date': order_date,
                        'order_number': order_number,
                        'order_company': order_company,
                        'order_division': order_division,
                        'order_launch_date': order_launch_date,
                        'order_execution_date': order_execution_date,
                        'responsible': responsible,
                        'comment': comment,
                    }

    pprint(orders_data['202300920'])

    # Создаем таблицы в БД
    engine = create_engine('sqlite:///data/orders_row_data.db')
    Base.metadata.drop_all(engine)
    Base.metadata.create_all(engine)
    session = Session(bind=engine)

    # Импортируем данные в БД
    session = import_data_to_db_production_orders(orders_data=orders_data,
                                                  session=session)
    # Сохраняем таблицы в БД
    session.commit()
