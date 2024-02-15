"""
    Скрипт имитирует работу коннектора программы 1С для экспорта
    данных в БД.
    На данном этапе данные из программы 1С, вручную
    копируются в файл excel на отдельные листы, предназначенные
    для каждого вида отчета.
    Скрипт последовательно проходит по листам excel и собирает
    данные по заказу на производство в таблицы данных на основе классов.
    После сбора всей информации с листов, скрипт импортирует
    данные из словаря в таблицы БД.
"""

from datetime import datetime

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from app import (ArticleExtractor, Base, ReportSettingsOrders,
                 extract_data_contractors,
                 extract_data_to_report_monitor_for_work_centers,
                 extract_data_to_report_moving_sets_of_furnuture,
                 extract_data_to_report_production_orders_report,
                 extract_data_to_report_release_of_assembly_kits,
                 import_data_to_db_main_orders)

config = ReportSettingsOrders()
extractor = ArticleExtractor()
orders_data = dict()
error_log = dict()

if __name__ == '__main__':
    # Собираем данные о контрагентах
    path_to_file = config.path_dir_local + config.path_data + config.file_name_orders2 + config.macros
    wb = load_workbook(path_to_file, read_only=True, keep_vba=True, data_only=True, keep_links=False)
    sheet_name = config.sheet_contractors
    workbook = wb[sheet_name]
    contractors = dict()
    contractors = extract_data_contractors(
        contractors=contractors,
        workbook=workbook,
        config=config
        )

    # Открыть файл с исходными данными
    print("Открываем файл с исходными данными")
    start = datetime.now()
    workbook = load_workbook(
        filename=config.source_file,
        read_only=True,
        data_only=True
    )
    end = datetime.now()
    total_time = (end - start).total_seconds()
    print(f"Время на открытие файла: {total_time} с")

    # Собираем данные по перемещению комплектов мебели
    orders_data = extract_data_to_report_moving_sets_of_furnuture(
        orders_data=orders_data,
        contractors=contractors,
        workbook=workbook,
        sheet=config.sheet_moving_1C,
        expression=config.expression)

    # Собираем данные по выпуску комплектов мебели
    orders_data = extract_data_to_report_release_of_assembly_kits(
        orders_data=orders_data,
        workbook=workbook,
        sheet=config.sheet_kits_1C,
        expression=config.expression)

    # Собираем данные по проценту готовности заказов
    orders_data = extract_data_to_report_monitor_for_work_centers(
        orders_data=orders_data,
        workbook=workbook,
        sheet=config.sheet_percentage_1C)

    # Собираем основной и дополнительные заказы с их описанием
    orders_data = extract_data_to_report_production_orders_report(
        orders_data=orders_data,
        workbook=workbook,
        sheet=config.sheet_division_1C)

    # Создаем таблицы в БД
    engine = create_engine('sqlite:///data/orders_row_data.db')
    Base.metadata.drop_all(engine)
    Base.metadata.create_all(engine)
    session = Session(bind=engine)

    session = import_data_to_db_main_orders(
        orders_data=orders_data,
        session=session)

    start = datetime.now()
    session.commit()
    end = datetime.now()
    total_time = (end - start).total_seconds()
    print(f"Время на сохранение данных: {total_time} с")
    print("Программа завершена!")
