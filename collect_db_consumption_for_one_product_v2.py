import configparser
from datetime import datetime

from openpyxl import Workbook, load_workbook

from app import ArticleExtractor

config = configparser.ConfigParser()

config.read('settings.ini', encoding='utf8')

exl_file_dir = config['PATH.DIR']['Path_dir']
exl_data_dir = config['PATH.DIR']['Path_data']

file_name_orders1 = config['File.name']['File_name_orders1']
file_name_orders2 = config['File.name']['File_name_orders2']

file_name_source_db_1 = config['File.name']['File_name_source_db_1']
file_name_source_db_2 = config['File.name']['File_name_source_db_2']

sales_forecast = config['File.name']['Sales_forecast']

macros = config['File.extension']['macros']
not_macros = config['File.extension']['not_macros']

deploy = config['Sheet.name']['Deploy']
forecast = config['Sheet.name']['Forecast']
not_deploy = config['Sheet.name']['Not_deploy']
sheet_main = config['Sheet.name']['Sheet_main']
sheet_all_categories = config['Sheet.name']['Sheet_all_categories']
sheet_sheet_material = config['Sheet.name']['Sheet_sheet_material']

# Названия файлов с заказами
orders_files = [
    file_name_orders1,
    file_name_orders2,
]

# Данные по заказам на производство
order_data = dict()

# Извлечь шестизначный артикул из наименования мебели
extractor = ArticleExtractor()

# Путь до файла с прогнозом продаж на год
exl_file = exl_file_dir + exl_data_dir + sales_forecast + not_macros

# Данные о статусе мебели
furniture_status = dict()

workbook = load_workbook(
    filename=exl_file,
    read_only=True,
    data_only=True
)
sheet = workbook['Прогноз 2024']

# Собрать статусы мебели из прогноза
for value in sheet.iter_rows(min_row=6, max_col=3, values_only=True):
    furniture_name = value[1]
    status = value[2]
    if status:
        article = extractor.get_article(furniture_name)
        if not furniture_status.get(article):
            furniture_status[article] = status

# Собрать данные о количестве мебели в заказе
for _ in range(len(orders_files)):
    exl_file = exl_file_dir + exl_data_dir + orders_files[_] + macros
    workbook = load_workbook(
        filename=exl_file,
        read_only=True,
        data_only=True
        )

    sheet = workbook[sheet_main]

    for value in sheet.iter_rows(min_row=7,
                                 max_col=7,
                                 values_only=True):
        key = value[3]       # Полный номер заказа на производство
        name = value[4]      # Наименование изделия мебели
        ordered = value[5]   # Количество изделий в заказе
        article = extractor.get_article(name)   # Артикул изделия мебели
        if furniture_status.get(article):
            furn_status = furniture_status[article]
        else:
            furn_status = 'не определен'

        if not order_data.get(key):
            order_data[key] = {
                'furniture_name': name,
                'furniture_article': article,
                'furniture_status': furn_status,
                'ordered': ordered,
            }

# Названия листов с данными по материалам из выгрузки
sheets = [
    sheet_all_categories + " " + forecast,
    sheet_sheet_material + " " + forecast,
]

# Файлы из выгрузки складской программы с разбивкой на материалы
material_files = [
    file_name_source_db_1,
    file_name_source_db_2,
]

article_data = dict()      # Собрать данные по расходу материала на 1 изделие  
material_status = dict()   # Собрать данные о ствтусе для материалов
working_status = ['Рабочий 2024', 'Под заказ', 'Новинка 2024', 'не определен']

# Собираем данные о расходе материалов на 1 изделие из листа Прогноз
for _ in range(len(material_files)):
    exl_file = exl_file_dir + exl_data_dir + material_files[_] + not_macros
    workbook = load_workbook(
        filename=exl_file,
        read_only=True,
        data_only=True,
    )
    for i in range(len(sheets)):
        sheet = workbook[sheets[i]]

        for value in sheet.iter_rows(min_row=2, values_only=True):
            key = value[1]                # Полный номер заказа на производство
            material_code = value[3]      # Код материала
            material_code.replace(" ", "")
            material_article = value[4]   # Артикул материала
            material_name = value[5]      # Наименованме материала
            # Количество материала в заказе
            material_amount = value[7] if value[7] else 0

            if order_data.get(key):
                furniture_name = order_data[key]['furniture_name']
                furniture_article = order_data[key]['furniture_article']
                furn_status = order_data[key]['furniture_status']
                if material_status.get(material_code):
                    if furn_status in working_status:
                        mat_status = 'не уникальный'
                        material_status[material_code] = mat_status
                else:
                    if furn_status in working_status:
                        mat_status = 'не уникальный'
                        material_status[material_code] = mat_status
                    else:
                        mat_status = 'уникальный'
                        material_status[material_code] = mat_status
                # Расход материала из разбивки складской программы
                cons_temp = order_data[key]['ordered']
                consumption_per_order = cons_temp if cons_temp else 0
                try:
                    consumption_per_1_product = material_amount / consumption_per_order
                except ZeroDivisionError:
                    consumption_per_1_product = 0
                # Добавить данные по артикулу и материалу
                if not article_data.get(furniture_article):
                    material = {
                        'furniture_name': furniture_name,
                        'furniture_status': furn_status,
                        material_code: {
                            'material_article': material_article,
                            'material_name': material_name,
                            'consumption_per_1_product': consumption_per_1_product,
                            }
                        }
                    article_data[furniture_article] = material
                # Если артикул есть в базе, добавить материал
                if article_data.get(furniture_article) and not article_data[furniture_article].get(material_code):
                    article_data[furniture_article][material_code] = {
                        'material_article': material_article,
                        'material_name': material_name,
                        'consumption_per_1_product': consumption_per_1_product,
                    }

# Сбор данных из оставшихся категорий
# Названия листов с данными по материалам из выгрузки
sheets = [
    sheet_all_categories + " " + not_deploy,
    sheet_all_categories + " " + deploy,
    sheet_sheet_material + " " + not_deploy,
    sheet_sheet_material + " " + deploy,
]

# Собираем данные о расходе материалов на 1 изделие из оставшихся листов
# для дополнения отсутствующей информации в прогнозе
for _ in range(len(material_files)):
    exl_file = exl_file_dir + exl_data_dir + material_files[_] + not_macros
    workbook = load_workbook(
        filename=exl_file,
        read_only=True,
        data_only=True,
    )
    for i in range(len(sheets)):
        sheet = workbook[sheets[i]]

        for value in sheet.iter_rows(min_row=2, values_only=True):
            key = value[1]                # Полный номер заказа на производство
            material_code = value[3]      # Код материала
            material_article = value[4]   # Артикул материала
            material_name = value[5]      # Наименованме материала
            # # Количество материала в заказе
            material_amount = value[7] if value[7] else 0

            if order_data.get(key):
                furniture_name = order_data[key]['furniture_name']
                furniture_article = order_data[key]['furniture_article']
                furn_status = order_data[key]['furniture_status']
                if material_status.get(material_code):
                    if furn_status in working_status:
                        mat_status = 'не уникальный'
                        material_status[material_code] = mat_status
                else:
                    if furn_status in working_status:
                        mat_status = 'не уникальный'
                        material_status[material_code] = mat_status
                    else:
                        mat_status = 'уникальный'
                        material_status[material_code] = mat_status
                cons_temp = order_data[key]['ordered']
                consumption_per_order = cons_temp if cons_temp else 0
                try:
                    consumption_per_1_product = material_amount / consumption_per_order
                except ZeroDivisionError:
                    consumption_per_1_product = 0
                if not article_data.get(furniture_article):
                    material = {
                        'furniture_name': furniture_name,
                        'furniture_status': furn_status,
                        material_code: {
                            'material_article': material_article,
                            'material_name': material_name,
                            'consumption_per_1_product': consumption_per_1_product,
                            }
                        }
                    article_data[furniture_article] = material
                if article_data.get(furniture_article) and not article_data[furniture_article].get(material_code):
                    article_data[furniture_article][material_code] = {
                            'material_article': material_article,
                            'material_name': material_name,
                            'consumption_per_1_product': consumption_per_1_product,
                        }

# Файл с результатами расчетов расхода на одно изделие
workbook = Workbook()

# Дата формирования отчета
dt_obj = datetime.now()
dt_str = dt_obj.strftime("%d.%m.%Y %Hч%Mм")

# Создаем лист отчета
sheet = workbook.active
sheet.title = 'Расход на 1 изделие'

# Названия колонок отчета
sheet.append([
    "Артикул",
    "Наименование",
    "Статус",
    "Код материала",
    "Артикул материала",
    "Наименование материала",
    "Уникальность материала",
    "Расход на 1 изделие",
    ])

# запись данных в файл отчета
for article in article_data:
    for mat_cod in article_data[article]:
        if mat_cod != 'furniture_name':
            if mat_cod != 'furniture_status':
                data = [
                    article,
                    article_data[article]['furniture_name'],
                    article_data[article]['furniture_status'],
                    mat_cod,
                    article_data[article][mat_cod]['material_article'],
                    article_data[article][mat_cod]['material_name'],
                    material_status[mat_cod],
                    article_data[article][mat_cod]['consumption_per_1_product'],
                ]
                sheet.append(data)

# Сохраняем данные в файл Excel
workbook.save(filename=exl_file_dir + exl_data_dir + sheet.title + ' от ' + dt_str + not_macros)
