"""Скрипт генерирует отчет о состоянии заказов на производство на
текущую дату в файл excel.

Источником данных является БД в которую предварительно импортируют
исходные данные из 1С Производство.

Отчет является вспомогательным источником данных для монитора заказов
на производство предназначенным для краткосрочного планирования потребности
к запуску заказов на производство.

Определяет готовность к сборке изделий и формирует задание на участок
сборки изделий.

Определяет возможность к запуску заказов в производство на основании
приоритета и технической готовности к запуску в цех раскрой и формирует
задание в цех раскрой.
"""

from datetime import datetime

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from sqlalchemy import create_engine
from sqlalchemy.orm import Session
from tqdm import tqdm

from app import (COLUMNS_FOR_INSERTING_FORMULAS_SUBTOTAL,
                 COLUMNS_FORMAT_PERCENTAGE, COLUMNS_WIDTH,
                 ORDERS_REPORT_COLUMNS_NAME, ORDERS_REPORT_MAIN_SHEET, Base,
                 DescriptionMainOrderRowData, MonitorForWorkCenters,
                 OrderRowData, ReleaseOfAssemblyKitsRowData,
                 ReportSettingsOrders, SubOrder, SubOrderReportDescription,
                 calc_number_products_cutting_and_painting_workshops,
                 calculate_percentage_of_painting_readiness,
                 calculation_number_details_fact_paint_to_assembly,
                 calculation_percentage_of_assembly,
                 determine_if_there_is_a_painting,
                 determine_ready_status_of_assembly, set_format_to_cell,
                 set_formula_to_cell, set_styles_to_cells, set_weigth_to_cell)
from app.styles_for_tables.orders_report import t_head, td_digit, td_text

# Обращаемся к БД Исходные данные по заказам на производство
start_all = datetime.now()
start = datetime.now()
print('Открываем БД исходные данные для отчета')
try:
    engine = create_engine('sqlite:///data/orders_row_data.db')
    session = Session(autoflush=False, bind=engine)
    Base.metadata.create_all(bind=engine)
except FileNotFoundError:
    print('По указанному пути файл не обнаружен')
end = datetime.now()
total_time = (end - start).total_seconds()
print(f'Выполнено за: {total_time} c.')

start = datetime.now()
print('Запрос в БД основной заказ')
with session as db:
    orders_report = db.query(OrderRowData).all()
end = datetime.now()
total_time = (end - start).total_seconds()
print(f'Выполнен за : {total_time} c.')

workbook = Workbook()
config = ReportSettingsOrders()

sheet = workbook.active
sheet.title = ORDERS_REPORT_MAIN_SHEET
sheet.append(ORDERS_REPORT_COLUMNS_NAME)

# Определить статус изделия: крашеное/не крашеное и наличие корпуса
product_is_painted = dict()
product_is_painted = determine_if_there_is_a_painting(
    orders_report=orders_report,
    product_is_painted=product_is_painted,
    session_db=db,
    table_in_db=SubOrderReportDescription   # Подчиненные заказы
)

# Запомнить номера колонок
colums_number = dict()
for number, name in enumerate(ORDERS_REPORT_COLUMNS_NAME, 1):
    colums_number[name] = number

# Записываем данные в файл отчета Excel
for order in tqdm(orders_report,
                  ncols=80,
                  ascii=True,
                  desc="Формирование отчета"):
    furniture_article = order.furniture_article
    ordered = order.ordered
    released = order.released
    paint_to_paint = ""
    cutting_status = ""
    order_division = ""
    order_launch_date = ""
    order_execution_date = ""
    responsible = ""
    comment = ""
    contractor = order.furniture_contractor

    # Описание основного заказа
    description_main = \
        db.query(DescriptionMainOrderRowData).filter(DescriptionMainOrderRowData.order_id == order.id)
    if description_main:
        for d in description_main:
            order_date = d.order_date
            order_number = d.order_number
            order_division = d.order_division
            order_launch_date = d.order_launch_date
            order_execution_date = d.order_execution_date
            responsible = d.responsible
            comment = d.comment
    cutting_shop_for_assembly = 0
    cutting_shop_for_painting = 0
    paint_shop_for_assembly = 0
    assembly_shop = 0

    # Выпуск комплектов мебели (отчет матсеров)
    release_of_assembly = \
        db.query(ReleaseOfAssemblyKitsRowData).filter(ReleaseOfAssemblyKitsRowData.order_id == order.id)
    if release_of_assembly:
        for r in release_of_assembly:
            cutting_shop_for_assembly = r.cutting_shop_for_assembly
            cutting_shop_for_painting = r.cutting_shop_for_painting
            paint_shop_for_assembly = r.paint_shop_for_assembly
            assembly_shop = r.assembly_shop

    # Процент готовности сборка
    percentg_of_assembly = calculation_percentage_of_assembly(
        ordered=ordered,
        released=released,
        assembly_shop=assembly_shop
    )

    # Данные монитора рабочих центров для основного заказа
    percentage_of_readiness_to_cut = "не развернут"
    number_of_details_plan = 0
    number_of_details_fact = 0

    # Монитор рабочих центров (основной заказ) (готовность раскрой на буфер и раскрой на покраску)
    monitor = \
        db.query(MonitorForWorkCenters).filter(MonitorForWorkCenters.order_id == order.id)
    if monitor:
        for m in monitor:
            percentage_of_readiness_to_cut = m.percentage_of_readiness_to_cut
            number_of_details_plan = m.number_of_details_plan
            number_of_details_fact = m.number_of_details_fact

    # Монитор рабочих центров (подчиненные заказы)
    monitor_sub = db.query(SubOrder).filter(SubOrder.order_id == order.id)

    # Данные для подчиненных заказов (раскрой на буфер)
    composite_key_cut_to_assembly: str = ""
    percentage_of_readiness_to_cut_to_assembly: float = 0
    number_of_details_plan_cut_to_assembly: int = 0
    number_of_details_fact_cut_to_assembly: int = 0
    type_of_movement_cut_to_assembly: str = ""

    # Данные для подичиненных заказов (раскрой на покраску)
    composite_key_cut_to_paint: str = ""
    percentage_of_readiness_to_cut_to_paint: float = 0
    number_of_details_plan_cut_to_paint: int = 0
    number_of_details_fact_cut_to_paint: int = 0
    type_of_movement_cut_to_paint: str = ""
    if monitor_sub:
        for m_sub in monitor_sub:
            if m_sub.type_of_movement == "Раскрой на буфер":
                composite_key_cut_to_assembly = m_sub.composite_key
                percentage_of_readiness_to_cut_to_assembly = m_sub.percentage_of_readiness_to_cut
                number_of_details_plan_cut_to_assembly = m_sub.number_of_details_plan
                number_of_details_fact_cut_to_assembly = m_sub.number_of_details_fact
                type_of_movement_cut_to_assembly = m_sub.type_of_movement
            if m_sub.type_of_movement == "Раскрой на покраску":
                composite_key_cut_to_paint = m_sub.composite_key
                percentage_of_readiness_to_cut_to_paint = m_sub.percentage_of_readiness_to_cut
                number_of_details_plan_cut_to_paint = m_sub.number_of_details_plan
                number_of_details_fact_cut_to_paint = m_sub.number_of_details_fact
                type_of_movement_cut_to_paint = m_sub.type_of_movement

    # Статус изделия: (крашеное/не крашеное) (есть корпус/нет корпуса)
    painted_status = "не определено"
    cutting_status = "не определено"
    if product_is_painted.get(furniture_article):
        painted_status = product_is_painted[furniture_article]['painted_status']
        cutting_status = product_is_painted[furniture_article]['cutting_status']

    # Расчет процента готовности покраски
    percentage_of_readiness_painting = calculate_percentage_of_painting_readiness(
        painted_status=painted_status,
        cutting_shop_for_painting=cutting_shop_for_painting,
        paint_shop_for_assembly=paint_shop_for_assembly,
    )

    # Описание подчиненных заказов (раскрой на буфер, раскрой на покраску и покраска на буфер)
    description_sub = \
        db.query(SubOrderReportDescription).filter(SubOrderReportDescription.order_id == order.id)
    order_division_paint = ""
    composite_key_paint = ""
    number_of_details_fact_paint_to_assembly = ""
    if description_sub:
        for d_sub in description_sub:
            if d_sub.order_division == "Цех покраски":
                order_division_paint = d_sub.order_division
                composite_key_paint = d_sub.composite_key
                # Расчет количества деталей покраски на буфер (расчетный показатель)
                number_of_details_fact_paint_to_assembly = \
                    calculation_number_details_fact_paint_to_assembly(
                        percentage_of_readiness_painting=percentage_of_readiness_painting,
                        number_of_details_plan_cut_to_paint=number_of_details_plan_cut_to_paint
                        )
    assembly_ready_status = ""
    quantity_to_be_assembled = ""
    cut_to_the_buffer_in_progress = ""
    cutting_for_painting_in_progress = ""
    painting_in_progress = ""

    # Итоговые данные отчета
    data = [
        furniture_article,
        order.composite_key,
        order_date,
        order_number,
        order.full_order_number,
        order.furniture_name,
        ordered,
        released,
        order.remains_to_release,
        cutting_shop_for_assembly,
        cutting_shop_for_painting,
        paint_shop_for_assembly,
        assembly_shop,
        painted_status,
        cutting_status,
        percentg_of_assembly,
        percentage_of_readiness_to_cut,
        percentage_of_readiness_painting,
        comment,
        contractor,
        order_division,
        order_launch_date,
        order_execution_date,
        assembly_ready_status,
        quantity_to_be_assembled,
        cut_to_the_buffer_in_progress,
        cutting_for_painting_in_progress,
        painting_in_progress,
        number_of_details_plan,
        number_of_details_fact,
        responsible,
        type_of_movement_cut_to_assembly,
        composite_key_cut_to_assembly,
        percentage_of_readiness_to_cut_to_assembly,
        number_of_details_plan_cut_to_assembly,
        number_of_details_fact_cut_to_assembly,
        type_of_movement_cut_to_paint,
        composite_key_cut_to_paint,
        percentage_of_readiness_to_cut_to_paint,
        number_of_details_plan_cut_to_paint,
        number_of_details_fact_cut_to_paint,
        order_division_paint,
        composite_key_paint,
        percentage_of_readiness_painting,
        number_of_details_plan_cut_to_paint,
        number_of_details_fact_paint_to_assembly
    ]
    sheet.append(data)

# Обращаеся к основному листу
worksheet = workbook[ORDERS_REPORT_MAIN_SHEET]

# Вставка строки
worksheet.insert_rows(2, 3)

# Вставка формул
start_row = 5
formula_row = 3
auto_filter = 4
last_row = db.query(OrderRowData.id).count() + auto_filter

# Установить формулу итогов в шапку таблицы
set_formula = set_formula_to_cell(
    formula="SUBTOTAL",
    formula_param=[9],
    worksheet=worksheet,
    row=formula_row,
    start_row=start_row,
    last_row=last_row,
    columns_number=colums_number,
    columns_with_formula=COLUMNS_FOR_INSERTING_FORMULAS_SUBTOTAL,
    converter_letter=get_column_letter,
)

# Вставка автофильтра
last_coll = len(ORDERS_REPORT_COLUMNS_NAME)
column_last_letter = get_column_letter(last_coll)
worksheet.auto_filter.ref = f"A{auto_filter}:{column_last_letter}{last_coll}"

# Сортировка по возрастанию для колонки Дата (от старых к новым)
column_letter = get_column_letter(colums_number['Дата'])
worksheet.auto_filter.add_sort_condition(
    f"{column_letter}{start_row}:{column_letter}{last_row}",
    descending=False
    )

# Закрепление строки
worksheet.freeze_panes = f"A{start_row}"

# Регистрация стилей
workbook.add_named_style(t_head)
workbook.add_named_style(td_digit)
workbook.add_named_style(td_text)

# Заливка, выравнивание, границы шапки таблицы
for i in range(last_coll):
    column_letter = get_column_letter(i + 1)
    worksheet[f'{column_letter}1'].style = 'table_header'
    worksheet[f'{column_letter}2'].style = 'table_header'
    worksheet[f'{column_letter}{formula_row}'].style = 'table_header'
    worksheet[f'{column_letter}{formula_row}'].font = Font(
        name='Times New Roma',
        size=11,
        bold=True,
        )
    worksheet[f'{column_letter}{auto_filter}'].fill = PatternFill('solid', fgColor="00FF9900")
    worksheet[f'{column_letter}{auto_filter}'].border = Border(
        left=Side(border_style='thin', color='FF000000'),
        right=Side(border_style='thin', color='FF000000'),
        top=Side(border_style='thin', color='FF000000'),
        bottom=Side(border_style='thin', color='FF000000'),
        )

# Стиль для ячеек
set_styles = set_styles_to_cells(
    worksheet=worksheet,
    start_row=start_row,
    last_row=last_row,
    column_name_left='Артикул',
    column_name_right='Количество деталей факт, покраска на буфер',
    columns_number=colums_number,
    converter_letter=get_column_letter,
    styles='table_data_text',
)

set_styles = set_styles_to_cells(
    worksheet=worksheet,
    start_row=start_row,
    last_row=last_row,
    column_name_left='Заказано',
    column_name_right='Выпущено сборка, данные мастеров',
    columns_number=colums_number,
    converter_letter=get_column_letter,
    styles='table_data_digit',
)

set_styles = set_styles_to_cells(
    worksheet=worksheet,
    start_row=start_row,
    last_row=last_row,
    column_name_left='Процент готовности сборка',
    column_name_right='Процент готовности покраска',
    columns_number=colums_number,
    converter_letter=get_column_letter,
    styles='table_data_digit',
)

set_styles = set_styles_to_cells(
    worksheet=worksheet,
    start_row=start_row,
    last_row=last_row,
    column_name_left='Раскрой на буфер, в работе',
    column_name_right='Количество деталей факт',
    columns_number=colums_number,
    converter_letter=get_column_letter,
    styles='table_data_digit',
)

set_styles = set_styles_to_cells(
    worksheet=worksheet,
    start_row=start_row,
    last_row=last_row,
    column_name_left='Процент готовности, раскрой на буфер',
    column_name_right='Количество деталей факт, раскрой на буфер',
    columns_number=colums_number,
    converter_letter=get_column_letter,
    styles='table_data_digit',
)

set_styles = set_styles_to_cells(
    worksheet=worksheet,
    start_row=start_row,
    last_row=last_row,
    column_name_left='Процент готовности, раскрой на покраску',
    column_name_right='Количество деталей факт, раскрой на покраску',
    columns_number=colums_number,
    converter_letter=get_column_letter,
    styles='table_data_digit',
)

set_styles = set_styles_to_cells(
    worksheet=worksheet,
    start_row=start_row,
    last_row=last_row,
    column_name_left='Процент готовности, покраска на буфер',
    column_name_right='Количество деталей факт, покраска на буфер',
    columns_number=colums_number,
    converter_letter=get_column_letter,
    styles='table_data_digit',
)

# Ширина колонок отчета
set_weigth = set_weigth_to_cell(
    worksheet=worksheet,
    columns_number=colums_number,
    columns_with_format=COLUMNS_WIDTH,
    converter_letter=get_column_letter
)

green_color = '00339966'
yellow_color = '00FFFF00'
red_color = '00FF0000'

# Основные расчеты по заказам
for value in tqdm(worksheet.iter_rows(min_row=start_row, max_col=last_coll),
                  ncols=80,
                  ascii=True,
                  desc="Основные расчеты по заказам"):
    percentage = value[colums_number['Процент готовности сборка'] - 1].value
    if percentage >= 0.9:
        column_letter_left = get_column_letter(colums_number['Артикул'])
        column_letter_right = get_column_letter(colums_number['Наличие корпуса'])
        for cells in worksheet[f'{column_letter_left}{value[colums_number["Артикул"]].row}:{column_letter_right}{value[colums_number["Артикул"]].row}']:
            for cell in cells:
                cell.fill = PatternFill('solid', fgColor=green_color)
    if percentage < 0.9 and percentage > 0.1:
        column_letter_left = get_column_letter(colums_number['Заказано'])
        column_letter_right = get_column_letter(colums_number['Осталось выпустить'])
        for cells in worksheet[f'{column_letter_left}{value[colums_number["Артикул"]].row}:{column_letter_right}{value[colums_number["Артикул"]].row}']:
            for cell in cells:
                cell.fill = PatternFill('solid', fgColor=yellow_color)


        column_letter_left = get_column_letter(colums_number['Выпущено сборка, данные мастеров'])
        column_letter_right = get_column_letter(colums_number['Наличие корпуса'])
        for cells in worksheet[f'{column_letter_left}{value[colums_number["Артикул"]].row}:{column_letter_right}{value[colums_number["Артикул"]].row}']:
            for cell in cells:
                cell.fill = PatternFill('solid', fgColor=yellow_color)


    column_letter_left = get_column_letter(colums_number['Раскрой на буфер'])
    column_letter_right = get_column_letter(colums_number['Покраска на буфер'])
    for cells in worksheet[f'{column_letter_left}{value[colums_number["Артикул"]].row}:{column_letter_right}{value[colums_number["Артикул"]].row}']:
            for cell in cells:
                value_division = cell.value
                if value_division:
                    cell.fill = PatternFill('solid', fgColor=green_color)

    # Расчеты для колонок: (Статус готовности, сборка) и (Готово к сборке, количество)
    ordered = value[colums_number["Заказано"] - 1].value
    released = value[colums_number["Выпущено"] - 1].value
    remains_to_release = value[colums_number["Осталось выпустить"] - 1].value
    cutting_shop_for_assembly = value[colums_number['Раскрой на буфер'] - 1].value
    cutting_shop_for_painting = value[colums_number['Раскрой на покраску'] - 1].value
    paint_shop_for_assembly = value[colums_number['Покраска на буфер'] - 1].value
    painted_status = value[colums_number['Крашеное/не крашеное'] - 1].value
    cutting_status = value[colums_number['Наличие корпуса'] - 1].value
    percentg_of_assembly = value[colums_number['Процент готовности сборка'] - 1].value
    percentage_of_readiness_to_cut = value[colums_number["Процент готовности раскрой"] - 1].value
    type_of_movement_cut_to_assembly = value[colums_number['Тип перемещения деталей, раскрой на буфер'] - 1].value
    type_of_movement_cut_to_paint = value[colums_number['Тип перемещения деталей, раскрой на покраску'] - 1].value
    order_division_paint = value[colums_number['Тип перемещения деталей, покраска на буфер'] - 1].value
    
    # Определяем статус готовности изделия и количество к сборке
    assembly_ready_status, quantity_to_be_assembled = determine_ready_status_of_assembly(
        released=released,
        cutting_shop_for_assembly=cutting_shop_for_assembly,
        cutting_shop_for_painting=cutting_shop_for_painting,
        paint_shop_for_assembly=paint_shop_for_assembly,
        painted_status=painted_status,
        cutting_status=cutting_status,
        percentg_of_assembly=percentg_of_assembly,
        percentage_of_readiness_to_cut=percentage_of_readiness_to_cut,
    )

    # Записываем данные в ячейки
    value[colums_number["Статус готовности, сборка"] - 1].value = assembly_ready_status
    value[colums_number["Готово к сборке, количество"] - 1].value = quantity_to_be_assembled
    
    if quantity_to_be_assembled:
        value[colums_number["Готово к сборке, количество"] - 1].fill = PatternFill('solid', fgColor=green_color)

    cut_to_the_buffer_in_progress, \
        cutting_for_painting_in_progress, \
            painting_in_progress = calc_number_products_cutting_and_painting_workshops(
                ordered=ordered,
                released=released,
                remains_to_release=remains_to_release,
                cutting_shop_for_assembly=cutting_shop_for_assembly,
                cutting_shop_for_painting=cutting_shop_for_painting,
                paint_shop_for_assembly=paint_shop_for_assembly,
                painted_status=painted_status,
                cutting_status=cutting_status,
                assembly_ready_status=assembly_ready_status,
            )
    value[colums_number["Раскрой на буфер, в работе"] - 1].value = cut_to_the_buffer_in_progress
    if cut_to_the_buffer_in_progress:
        value[colums_number["Раскрой на буфер, в работе"] - 1].fill = PatternFill('solid', fgColor=green_color)
    value[colums_number["Раскрой на покраску, в работе"] - 1].value = cutting_for_painting_in_progress
    if cutting_for_painting_in_progress:
        value[colums_number["Раскрой на покраску, в работе"] - 1].fill = PatternFill('solid', fgColor=green_color)
    value[colums_number["Покраска, в работе"] - 1].value = painting_in_progress
    if painting_in_progress:
        value[colums_number["Покраска, в работе"] - 1].fill = PatternFill('solid', fgColor=green_color)

# Задать формат ячеек для колонок с процентом готовности
set_format = set_format_to_cell(
    format_cell="0.0%",
    worksheet=worksheet,
    start_row=start_row,
    last_row=last_row,
    columns_number=colums_number,
    columns_with_format=COLUMNS_FORMAT_PERCENTAGE,
    converter_letter=get_column_letter,
)

# Условное форматирование стобца с процентами
column_letter = get_column_letter(colums_number['Процент готовности сборка'])
worksheet.conditional_formatting.add(f'{column_letter}{start_row}:{column_letter}{last_row}',
                                     ColorScaleRule(start_type='percentile',
                                                    start_value=0,
                                                    start_color=red_color,
                                                    mid_type='percentile',
                                                    mid_value=50,
                                                    mid_color=yellow_color,
                                                    end_type='percentile',
                                                    end_value=100,
                                                    end_color=green_color)
                                                    )

# Условное форматирование стобца с процентами
column_letter = get_column_letter(colums_number['Процент готовности раскрой'])
worksheet.conditional_formatting.add(f'{column_letter}{start_row}:{column_letter}{last_row}',
                                     ColorScaleRule(start_type='percentile',
                                                    start_value=0,
                                                    start_color=red_color,
                                                    mid_type='percentile',
                                                    mid_value=50,
                                                    mid_color=yellow_color,
                                                    end_type='percentile',
                                                    end_value=100,
                                                    end_color=green_color)
                                                    )

# Условное форматирование стобца с процентами
column_letter = get_column_letter(colums_number['Процент готовности покраска'])
worksheet.conditional_formatting.add(f'{column_letter}{start_row}:{column_letter}{last_row}',
                                     ColorScaleRule(start_type='percentile',
                                                    start_value=0,
                                                    start_color=red_color,
                                                    mid_type='percentile',
                                                    mid_value=50,
                                                    mid_color=yellow_color,
                                                    end_type='percentile',
                                                    end_value=100,
                                                    end_color=green_color)
                                                    )

# Сохранение файла
dt = datetime.now()
dt = dt.strftime("%d.%m.%Y %Hч%Mм")
start = datetime.now()
print('Сохранение отчета в файл Excel')
workbook.save(filename=config.path_dir + config.report_file_name + " от " + dt + config.not_macros)
end = datetime.now()
total_time = (end - start).total_seconds()
print(f'Выполнено за : {total_time} c.')
total_time_all = (end - start_all).total_seconds()
print(f'Программа завершена за: {total_time_all} c')
